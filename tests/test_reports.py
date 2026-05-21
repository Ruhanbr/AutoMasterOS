"""
Tests for the reports module.

Covers:
  - WhatsApp link generation (various phone formats)
  - OS PDF generation returns bytes with %PDF header
  - Financial Excel generation returns valid xlsx bytes
"""
from decimal import Decimal
from datetime import datetime, timezone

import pytest

from app.modules.reports.whatsapp import build_os_whatsapp_message, build_whatsapp_link
from app.modules.reports.os_pdf import generate_os_pdf
from app.modules.reports.excel_generator import generate_financial_excel, generate_stock_excel


# ── Tests: WhatsApp link ──────────────────────────────────────────────────────

class TestBuildWhatsappLink:
    def test_plain_number_adds_ddi_55(self):
        link = build_whatsapp_link("11999990000", "Olá")
        assert link.startswith("https://wa.me/5511999990000")

    def test_already_has_ddi_55(self):
        link = build_whatsapp_link("5511999990000", "Teste")
        assert "wa.me/5511999990000" in link

    def test_with_plus_prefix(self):
        link = build_whatsapp_link("+5511999990000", "Teste")
        assert "wa.me/5511999990000" in link

    def test_formatted_phone_strips_non_digits(self):
        link = build_whatsapp_link("(11) 99999-0000", "Teste")
        assert "wa.me/5511999990000" in link

    def test_message_encoded_in_url(self):
        link = build_whatsapp_link("11999990000", "Olá Mundo!")
        assert "Ol%C3%A1" in link or "Ol" in link

    def test_short_number_gets_55_prefix(self):
        # 9-digit without DDD, should still get 55 prefix
        link = build_whatsapp_link("999990000", "Hi")
        assert link.startswith("https://wa.me/55999990000")

    def test_full_international_number_no_double_55(self):
        link = build_whatsapp_link("5521987654321", "Hi")
        assert link.startswith("https://wa.me/5521987654321")
        assert "5555" not in link


class TestBuildOsWhatsappMessage:
    def test_message_contains_required_fields(self):
        msg = build_os_whatsapp_message(
            client_name="João Silva",
            os_number=42,
            total="440,00",
            workshop_name="Oficina Teste",
        )
        assert "João Silva" in msg
        assert "#42" in msg
        assert "440,00" in msg
        assert "Oficina Teste" in msg

    def test_message_is_string(self):
        msg = build_os_whatsapp_message("Ana", 1, "100,00", "Workshop")
        assert isinstance(msg, str)


# ── Tests: OS PDF ─────────────────────────────────────────────────────────────

SAMPLE_ORDER_DATA = {
    "os_number": 42,
    "status": "FINALIZADA",
    "opened_at": datetime(2026, 4, 1, 8, 0, tzinfo=timezone.utc),
    "finished_at": datetime(2026, 4, 1, 17, 30, tzinfo=timezone.utc),
    "technician_name": "Carlos Mecânico",
    "description": "Revisão geral preventiva",
    "diagnosis": "Motor apresentando ruído anormal",
    "solution": "Troca de correia dentada e tensor",
    "client_name": "João Silva",
    "client_document": "123.456.789-01",
    "client_phone": "11999990000",
    "machine_model": "7200",
    "machine_brand": "John Deere",
    "machine_serial": "JD-ABC12345",
    "items": [
        {
            "item_type": "SERVICO",
            "description": "Mão de obra - Revisão geral",
            "quantity": Decimal("1.000"),
            "unit_price": Decimal("350.00"),
            "total_price": Decimal("350.00"),
        },
        {
            "item_type": "PECA",
            "description": "Filtro de óleo motor",
            "quantity": Decimal("2.000"),
            "unit_price": Decimal("45.00"),
            "total_price": Decimal("90.00"),
        },
    ],
    "total_services": Decimal("350.00"),
    "total_parts": Decimal("90.00"),
    "total_discount": Decimal("0.00"),
    "total_amount": Decimal("440.00"),
    "tenant_name": "Oficina Agrícola Teste",
    "generated_at": datetime.now(timezone.utc),
}


class TestGenerateOsPdf:
    def test_returns_bytes(self):
        result = generate_os_pdf(SAMPLE_ORDER_DATA)
        assert isinstance(result, bytes)

    def test_has_pdf_header(self):
        result = generate_os_pdf(SAMPLE_ORDER_DATA)
        assert result.startswith(b"%PDF")

    def test_non_empty_output(self):
        result = generate_os_pdf(SAMPLE_ORDER_DATA)
        assert len(result) > 1000  # PDF should be at least 1KB

    def test_without_machine_data(self):
        data = {**SAMPLE_ORDER_DATA, "machine_model": None, "machine_brand": None, "machine_serial": None}
        result = generate_os_pdf(data)
        assert result.startswith(b"%PDF")

    def test_without_diagnosis_solution(self):
        data = {**SAMPLE_ORDER_DATA, "diagnosis": None, "solution": None}
        result = generate_os_pdf(data)
        assert result.startswith(b"%PDF")

    def test_empty_items(self):
        data = {**SAMPLE_ORDER_DATA, "items": []}
        result = generate_os_pdf(data)
        assert result.startswith(b"%PDF")

    def test_aberta_status(self):
        data = {**SAMPLE_ORDER_DATA, "status": "ABERTA"}
        result = generate_os_pdf(data)
        assert result.startswith(b"%PDF")


# ── Tests: Financial Excel ────────────────────────────────────────────────────

SAMPLE_ENTRIES = [
    {
        "id": "aaaa-1111",
        "entry_type": "RECEITA",
        "description": "Receita da OS #1",
        "category": "Serviços e Peças",
        "amount": Decimal("440.00"),
        "reference_date": datetime(2026, 4, 1, tzinfo=timezone.utc),
        "service_order_id": "bbbb-2222",
        "notes": None,
        "created_at": datetime(2026, 4, 1, tzinfo=timezone.utc),
    },
    {
        "id": "cccc-3333",
        "entry_type": "DESPESA",
        "description": "Compra de ferramentas",
        "category": "Equipamentos",
        "amount": Decimal("200.00"),
        "reference_date": datetime(2026, 4, 2, tzinfo=timezone.utc),
        "service_order_id": None,
        "notes": "Nota fiscal 12345",
        "created_at": datetime(2026, 4, 2, tzinfo=timezone.utc),
    },
]

SAMPLE_SUMMARY = {
    "total_receitas": Decimal("440.00"),
    "total_despesas": Decimal("200.00"),
    "saldo": Decimal("240.00"),
    "date_from": None,
    "date_to": None,
}


class TestGenerateFinancialExcel:
    def test_returns_bytes(self):
        result = generate_financial_excel(SAMPLE_ENTRIES, SAMPLE_SUMMARY)
        assert isinstance(result, bytes)

    def test_valid_xlsx_header(self):
        """xlsx files start with the PK zip header."""
        result = generate_financial_excel(SAMPLE_ENTRIES, SAMPLE_SUMMARY)
        assert result[:2] == b"PK"

    def test_non_empty_output(self):
        result = generate_financial_excel(SAMPLE_ENTRIES, SAMPLE_SUMMARY)
        assert len(result) > 1000

    def test_empty_entries(self):
        result = generate_financial_excel([], {"total_receitas": 0, "total_despesas": 0, "saldo": 0})
        assert result[:2] == b"PK"

    def test_can_be_loaded_by_openpyxl(self):
        import io
        import openpyxl
        result = generate_financial_excel(SAMPLE_ENTRIES, SAMPLE_SUMMARY)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert "Lançamentos" in wb.sheetnames
        assert "Resumo" in wb.sheetnames


# ── Tests: Stock Excel ────────────────────────────────────────────────────────

SAMPLE_STOCK_ITEMS = [
    {
        "sku": "FILTRO-001",
        "description": "Filtro de óleo motor",
        "ncm_code": "84212300",
        "unit": "UN",
        "quantity": Decimal("10.000"),
        "min_quantity": Decimal("2.000"),
        "cost_price": Decimal("45.00"),
        "sale_price": Decimal("75.00"),
        "active": True,
        "created_at": datetime(2026, 4, 1, tzinfo=timezone.utc),
    },
    {
        "sku": "OLEO-15W40",
        "description": "Óleo 15W40 Mineral",
        "ncm_code": None,
        "unit": "LT",
        "quantity": Decimal("1.500"),  # below min_quantity
        "min_quantity": Decimal("5.000"),
        "cost_price": Decimal("12.50"),
        "sale_price": Decimal("20.00"),
        "active": True,
        "created_at": datetime(2026, 4, 1, tzinfo=timezone.utc),
    },
]


class TestGenerateStockExcel:
    def test_returns_bytes(self):
        result = generate_stock_excel(SAMPLE_STOCK_ITEMS)
        assert isinstance(result, bytes)

    def test_valid_xlsx_header(self):
        result = generate_stock_excel(SAMPLE_STOCK_ITEMS)
        assert result[:2] == b"PK"

    def test_can_be_loaded_by_openpyxl(self):
        import io
        import openpyxl
        result = generate_stock_excel(SAMPLE_STOCK_ITEMS)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert "Estoque" in wb.sheetnames

    def test_empty_items(self):
        result = generate_stock_excel([])
        assert result[:2] == b"PK"
