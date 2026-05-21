"""
Excel report generator using openpyxl.

Provides:
  - generate_financial_excel: financial entries + summary sheet
  - generate_stock_excel: stock items report
"""
import io
from datetime import datetime
from decimal import Decimal

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# Color constants (ARGB)
_COLOR_HEADER = "1A5276"       # dark blue
_COLOR_SUBHEADER = "2E86C1"    # medium blue
_COLOR_ACCENT = "F39C12"       # amber
_COLOR_ROW_ALT = "EBF5FB"      # light blue
_COLOR_GREEN = "1E8449"
_COLOR_RED = "C0392B"
_COLOR_GRAY = "566573"
_COLOR_WHITE = "FFFFFF"


def _header_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _header_font(size: int = 11, bold: bool = True, color: str = "FFFFFF") -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=color)


def _body_font(size: int = 10, bold: bool = False, color: str = "000000") -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=color)


def _thin_border() -> Border:
    thin = Side(style="thin", color="D5D8DC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


def _right() -> Alignment:
    return Alignment(horizontal="right", vertical="center")


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center")


def _fmt_currency(value) -> float:
    try:
        return float(Decimal(str(value)))
    except Exception:
        return 0.0


def _fmt_datetime(dt) -> str:
    if dt is None:
        return ""
    if isinstance(dt, str):
        return dt
    try:
        return dt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(dt)


def _apply_header_row(ws, row: int, headers: list[str], col_widths: list[int]) -> None:
    """Apply styled header row starting at given row."""
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = _header_font()
        cell.fill = _header_fill(_COLOR_HEADER)
        cell.alignment = _center()
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def generate_financial_excel(entries: list[dict], summary: dict) -> bytes:
    """
    Generates an Excel workbook with two sheets:
      1. Lançamentos: all financial entries
      2. Resumo: summary totals

    entries: list of dicts with keys:
      id, entry_type, description, category, amount, reference_date,
      service_order_id, notes, created_at

    summary: dict with keys:
      total_receitas, total_despesas, saldo, date_from, date_to
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: Lançamentos ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Lançamentos"
    ws1.sheet_view.showGridLines = False

    # Title row
    ws1.merge_cells("A1:H1")
    title_cell = ws1["A1"]
    title_cell.value = "RELATÓRIO FINANCEIRO — LANÇAMENTOS"
    title_cell.font = _header_font(size=14)
    title_cell.fill = _header_fill(_COLOR_HEADER)
    title_cell.alignment = _center()
    ws1.row_dimensions[1].height = 30

    # Generated at
    ws1.merge_cells("A2:H2")
    gen_cell = ws1["A2"]
    gen_cell.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    gen_cell.font = _body_font(size=9, color=_COLOR_GRAY)
    gen_cell.alignment = _right()
    ws1.row_dimensions[2].height = 18

    # Headers
    headers = ["Tipo", "Descrição", "Categoria", "Valor (R$)", "Data Referência", "OS Vinculada", "Obs.", "Criado em"]
    widths = [14, 40, 18, 16, 20, 36, 30, 20]
    _apply_header_row(ws1, 3, headers, widths)
    ws1.row_dimensions[3].height = 22

    # Data rows
    for row_idx, entry in enumerate(entries, start=4):
        alt = row_idx % 2 == 0
        row_fill = PatternFill("solid", fgColor=_COLOR_ROW_ALT) if alt else PatternFill("solid", fgColor=_COLOR_WHITE)

        entry_type = entry.get("entry_type", "")
        amount = _fmt_currency(entry.get("amount", 0))

        # Type cell
        c_type = ws1.cell(row=row_idx, column=1, value=entry_type)
        c_type.font = _body_font(
            bold=True,
            color=_COLOR_GREEN if entry_type == "RECEITA" else (_COLOR_RED if entry_type == "DESPESA" else _COLOR_GRAY),
        )
        c_type.alignment = _center()
        c_type.border = _thin_border()
        c_type.fill = row_fill

        cols = [
            entry.get("description", ""),
            entry.get("category", ""),
            amount,
            _fmt_datetime(entry.get("reference_date")),
            str(entry.get("service_order_id", "") or ""),
            entry.get("notes", "") or "",
            _fmt_datetime(entry.get("created_at")),
        ]
        for col_offset, val in enumerate(cols, start=2):
            cell = ws1.cell(row=row_idx, column=col_offset, value=val)
            cell.font = _body_font()
            cell.border = _thin_border()
            cell.fill = row_fill
            if col_offset == 4:  # amount column
                cell.alignment = _right()
                cell.number_format = '#,##0.00'
            else:
                cell.alignment = _left()

        ws1.row_dimensions[row_idx].height = 18

    ws1.freeze_panes = "A4"

    # ── Sheet 2: Resumo ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumo")
    ws2.sheet_view.showGridLines = False

    # Title
    ws2.merge_cells("A1:C1")
    t = ws2["A1"]
    t.value = "RESUMO FINANCEIRO"
    t.font = _header_font(size=14)
    t.fill = _header_fill(_COLOR_HEADER)
    t.alignment = _center()
    ws2.row_dimensions[1].height = 30

    summary_rows = [
        ("Total Receitas (R$)", _fmt_currency(summary.get("total_receitas", 0)), _COLOR_GREEN),
        ("Total Despesas (R$)", _fmt_currency(summary.get("total_despesas", 0)), _COLOR_RED),
        ("Saldo (R$)", _fmt_currency(summary.get("saldo", 0)), _COLOR_HEADER),
    ]
    for row_idx, (label, value, color) in enumerate(summary_rows, start=2):
        c_label = ws2.cell(row=row_idx, column=1, value=label)
        c_label.font = _body_font(bold=True)
        c_label.fill = _header_fill("EBF5FB")
        c_label.alignment = _left()
        c_label.border = _thin_border()
        ws2.column_dimensions["A"].width = 25

        c_value = ws2.cell(row=row_idx, column=2, value=value)
        c_value.font = Font(name="Calibri", size=11, bold=True, color=color)
        c_value.number_format = '#,##0.00'
        c_value.alignment = _right()
        c_value.border = _thin_border()
        ws2.column_dimensions["B"].width = 20
        ws2.row_dimensions[row_idx].height = 22

    if summary.get("date_from") or summary.get("date_to"):
        period_label = (
            f"Período: {_fmt_datetime(summary.get('date_from'))} a "
            f"{_fmt_datetime(summary.get('date_to'))}"
        )
        ws2.merge_cells("A6:C6")
        p = ws2["A6"]
        p.value = period_label
        p.font = _body_font(size=9, color=_COLOR_GRAY)
        p.alignment = _left()

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_stock_excel(items: list[dict]) -> bytes:
    """
    Generates a stock items Excel report.

    items: list of dicts with keys:
      sku, description, ncm_code, unit, quantity, min_quantity,
      cost_price, sale_price, active, created_at
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estoque"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = "RELATÓRIO DE ESTOQUE"
    t.font = _header_font(size=14)
    t.fill = _header_fill(_COLOR_HEADER)
    t.alignment = _center()
    ws.row_dimensions[1].height = 30

    # Generated at
    ws.merge_cells("A2:J2")
    gen = ws["A2"]
    gen.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    gen.font = _body_font(size=9, color=_COLOR_GRAY)
    gen.alignment = _right()
    ws.row_dimensions[2].height = 18

    headers = ["SKU", "Descrição", "NCM", "Unidade", "Qtd Atual", "Qtd Mínima", "Custo (R$)", "Venda (R$)", "Ativo", "Cadastrado em"]
    widths = [15, 40, 12, 10, 14, 14, 16, 16, 10, 20]
    _apply_header_row(ws, 3, headers, widths)
    ws.row_dimensions[3].height = 22

    for row_idx, item in enumerate(items, start=4):
        alt = row_idx % 2 == 0
        row_fill = PatternFill("solid", fgColor=_COLOR_ROW_ALT) if alt else PatternFill("solid", fgColor=_COLOR_WHITE)

        quantity = _fmt_currency(item.get("quantity", 0))
        min_qty = _fmt_currency(item.get("min_quantity", 0))
        is_low_stock = quantity < min_qty and min_qty > 0

        row_data = [
            item.get("sku", ""),
            item.get("description", ""),
            item.get("ncm_code", "") or "",
            item.get("unit", "UN"),
            quantity,
            min_qty,
            _fmt_currency(item.get("cost_price", 0)),
            _fmt_currency(item.get("sale_price", 0)),
            "Sim" if item.get("active") else "Não",
            _fmt_datetime(item.get("created_at")),
        ]

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _body_font()
            cell.border = _thin_border()
            cell.fill = row_fill

            # Numeric formatting
            if col_idx in {5, 6}:  # quantities
                cell.number_format = '#,##0.000'
                cell.alignment = _right()
                if col_idx == 5 and is_low_stock:
                    cell.font = Font(name="Calibri", size=10, bold=True, color=_COLOR_RED)
            elif col_idx in {7, 8}:  # prices
                cell.number_format = '#,##0.00'
                cell.alignment = _right()
            elif col_idx == 9:  # active
                cell.alignment = _center()
                cell.font = Font(
                    name="Calibri", size=10, bold=True,
                    color=_COLOR_GREEN if item.get("active") else _COLOR_RED,
                )
            else:
                cell.alignment = _left()

        ws.row_dimensions[row_idx].height = 18

    # Summary row
    total_row = len(items) + 4
    ws.merge_cells(f"A{total_row}:D{total_row}")
    summary_label = ws[f"A{total_row}"]
    summary_label.value = f"Total de itens: {len(items)}"
    summary_label.font = _body_font(bold=True)
    summary_label.alignment = _left()
    summary_label.fill = _header_fill("EBF5FB")

    ws.freeze_panes = "A4"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
